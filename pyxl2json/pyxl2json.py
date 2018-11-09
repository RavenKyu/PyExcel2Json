#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod: PyExcel2Json
====================================
.. moduleauthor:: Raven Lim <hong18s@gmail.com>
.. note::
"""
import openpyxl
import json
import ntpath

################################################################################
def excel_column_calculate(n:int, d:list):
    """
    Excel의 `A`, `AA` 와 같은 값을 Column 값을 쉽게 구할 수 있습니다.
    :param n: 첫 번째 Column 기준으로 1 이상의 수
    :param d: 빈 list
    :return: str
    >>> excel_column_calculate(1, []))
    'A'
    >>> excel_column_calculate(27, [])
    'AA'
    >>> excel_column_calculate(1000, [])
    'ALL'
    """
    if not n:
        return ''.join(d)
    n -= 1
    r = n // 26
    m = (n % 26)
    d.insert(0, chr(65 + m))
    return excel_column_calculate(r, d)

################################################################################
def _open_excel_file(filename) -> openpyxl.Workbook:
    """
    파라메터 filename의 위치에 있는 엑셀파일을 열어서
    openpyxl.Workbook 객체를 반환
    :param filename: .xlsx 형태의 엑셀파일
    :return: <openpyxl.Workbook>
    """
    return openpyxl.load_workbook(filename)

################################################################################
def _read_sheet(wb:openpyxl.Workbook, sheet:str):
    """
    파라메터 sheet를 Workbook에서 찾아서 반환
    :param wb: <openpyxl.Workbook>
    :param sheet: WorkSheet 이름
    :return: <openpyxl.Worksheet>
    """
    return wb[sheet]

################################################################################
def _read_head(ws:openpyxl.worksheet.Worksheet, data_range):
    # 헤드 범위가 지정되지 않았을 경우
    # 첫 번째 줄의 시작과 끝의 위치를 가지고 시작
    if not data_range:
        max_column = excel_column_calculate(ws.max_column, [])
        data_range = "A1:{}1".format(max_column)
    return [x.value for x in ws[data_range] for x in x]

################################################################################
def _read_data(ws:openpyxl.worksheet.Worksheet, data_range):
    data = []
    if not data_range:
        max_column = excel_column_calculate(ws.max_column, [])
        data_range = "A2:{}{}".format(max_column, ws.max_row)
    for row in ws[data_range]:
        cells = []
        for cell in row:
            cells.append(cell.value)
        data.append(cells)
    return data

################################################################################
def _to_dict(head: list, data: list):
    r = []
    for d in data:
        r.append(dict(zip(head, d)))
    return r

################################################################################
def _to_json(data: list):
    return json.dumps(data, indent=4, ensure_ascii=False)

################################################################################
def _as_file(path:str, data):
    head, tail = ntpath.split(path)
    filename_ext = tail or ntpath.basename(head)
    filename, ext = ntpath.splitext(filename_ext)
    with open(filename + '.json', 'w') as f:
        f.write(data)

################################################################################
def main(argspec):
    for filename in argspec.excel_filename:
        wb = _open_excel_file(filename)
        ws = _read_sheet(wb, argspec.sheet)
        head = _read_head(ws, argspec.head)
        data = _read_data(ws, argspec.data)
        d = _to_dict(head, data)
        json_data = _to_json(d)
        if argspec.verbose:
            print(json_data)
        if argspec.asfile:
            _as_file(filename, json_data)

